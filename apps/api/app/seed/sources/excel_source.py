"""SAT + SNICE spreadsheet catalog source.

Activated automatically by the seed script when the workbooks are present.
The SAT ``c_FraccionArancelaria`` catalog supplies codes, NICOs and
descriptions; the SNICE export supplies the IGI rate per tariff code.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.models import TariffItem
from app.seed.sources.base import CatalogSource

logger = logging.getLogger(__name__)

# Only consumer electronics and computer peripherals are in scope (vision.md).
IN_SCOPE_CHAPTERS = {"84", "85"}

# Header aliases, because the published files rename columns between releases.
_CODE_HEADERS = ("fraccion", "fraccion arancelaria", "c_fraccionarancelaria", "codigo")
_NICO_HEADERS = ("nico", "c_nico", "numero de identificacion comercial")
_DESCRIPTION_HEADERS = ("descripcion", "descripcion nico", "texto")
_UNIT_HEADERS = ("umt", "unidad de medida", "unidad")
_RATE_HEADERS = ("igi", "arancel", "arancel igi", "tasa", "advalorem", "ad valorem")


class ExcelCatalogSource(CatalogSource):
    """Parses the official spreadsheets into domain models."""

    def __init__(self, sat_path: Path, snice_path: Path | None = None) -> None:
        self._sat_path = sat_path
        self._snice_path = snice_path

    @property
    def name(self) -> str:
        suffix = (
            f" + SNICE ({self._snice_path})"
            if self._has_snice()
            else " (sin aranceles SNICE)"
        )
        return f"Excel SAT ({self._sat_path}){suffix}"

    def is_available(self) -> bool:
        return self._sat_path.is_file()

    def load(self) -> list[TariffItem]:
        igi_by_code = self._load_igi_rates() if self._has_snice() else {}
        rows = _read_rows(self._sat_path)

        items: list[TariffItem] = []
        seen: set[tuple[str, str]] = set()
        for row in rows:
            code = _normalise_code(_pick(row, _CODE_HEADERS))
            if code is None or code[:2] not in IN_SCOPE_CHAPTERS:
                continue

            nico = _normalise_nico(_pick(row, _NICO_HEADERS))
            key = (code, nico)
            if key in seen:
                continue
            seen.add(key)

            description = str(_pick(row, _DESCRIPTION_HEADERS) or "").strip()
            if not description:
                continue

            unit = str(_pick(row, _UNIT_HEADERS) or "").strip() or "Pza"
            items.append(
                TariffItem(
                    tariff_code=code,
                    nico=nico,
                    description=description,
                    heading_description="",
                    chapter=code[:2],
                    unit_of_measure=unit,
                    igi_rate=igi_by_code.get(code, 0.0),
                    iva_rate=0.16,
                    is_active=True,
                )
            )

        logger.info("Parsed %d tariff items from %s", len(items), self._sat_path.name)
        return items

    # --- internals ---------------------------------------------------------

    def _has_snice(self) -> bool:
        return self._snice_path is not None and self._snice_path.is_file()

    def _load_igi_rates(self) -> dict[str, float]:
        """Map tariff code to its IGI rate expressed as a fraction."""
        rates: dict[str, float] = {}
        assert self._snice_path is not None  # guarded by _has_snice
        for row in _read_rows(self._snice_path):
            code = _normalise_code(_pick(row, _CODE_HEADERS))
            if code is None:
                continue
            rate = _normalise_rate(_pick(row, _RATE_HEADERS))
            if rate is not None:
                rates.setdefault(code, rate)
        logger.info("Loaded %d IGI rates from SNICE", len(rates))
        return rates


def _read_rows(path: Path) -> list[dict[str, Any]]:
    """Read the first worksheet as a list of header-keyed dictionaries."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        headers = [_normalise_header(cell) for cell in header_row]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def _normalise_header(value: Any) -> str:
    """Lowercase, strip accents and collapse spaces so aliases can match."""
    text = str(value or "").strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    stripped = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", stripped)


def _pick(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    """Return the first column whose header matches one of the aliases."""
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    # Fall back to a substring match for headers like "descripcion de la fraccion".
    for header, value in row.items():
        if value in (None, ""):
            continue
        if any(alias in header for alias in aliases):
            return value
    return None


def _normalise_code(value: Any) -> str | None:
    """Return an 8 digit tariff code, or None when the cell is unusable."""
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 8 else None


def _normalise_nico(value: Any) -> str:
    """Return a 2 digit NICO, defaulting to zeros."""
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return "00"
    return digits[-2:].zfill(2)


def _normalise_rate(value: Any) -> float | None:
    """Convert a published rate into a fraction, e.g. 15 or 15% become 0.15."""
    text = str(value or "").strip().replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    # Published rates are percentages; anything above 1 is treated as such.
    return round(number / 100, 4) if number > 1 else round(number, 4)
